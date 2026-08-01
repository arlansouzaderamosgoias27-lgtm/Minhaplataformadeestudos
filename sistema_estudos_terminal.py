#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SISTEMA DE ESTUDOS - PMGO / PMDF (versao terminal)
====================================================
Versao para usar no computador, via terminal (cmd/PowerShell/bash).
Guarda os dados em um banco SQLite local (estudos_pmgo_pmdf.db),
no mesmo formato/logica do app HTML (20 revisoes por topico).

COMO USAR:
    python sistema_estudos_terminal.py

DEPENDENCIAS (instalar so se for exportar Excel/PDF):
    pip install openpyxl reportlab

Se essas bibliotecas nao estiverem instaladas, o sistema ainda
funciona normalmente para cadastro e controle de estudos - apenas
as opcoes de exportacao vao avisar que a biblioteca esta faltando.
"""

import sqlite3
import os
import sys
from datetime import date, datetime

DB_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "estudos_pmgo_pmdf.db")
TOTAL_REVISOES = 20
MATERIAS_PADRAO = [
    "Direito Constitucional", "Direito Penal", "Direito Processual Penal",
    "Português", "Matemática", "Informática", "Legislação", "Atualidades"
]


# ==========================================================
# BANCO DE DADOS
# ==========================================================
def conectar():
    conn = sqlite3.connect(DB_FILE)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def iniciar_banco():
    conn = conectar()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS materias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT UNIQUE NOT NULL
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS topicos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            materia_id INTEGER NOT NULL,
            nome TEXT NOT NULL,
            data_cadastro TEXT NOT NULL,
            concluido INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY(materia_id) REFERENCES materias(id) ON DELETE CASCADE
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS revisoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            topico_id INTEGER NOT NULL,
            numero INTEGER NOT NULL,
            feita INTEGER NOT NULL DEFAULT 0,
            data_feita TEXT,
            FOREIGN KEY(topico_id) REFERENCES topicos(id) ON DELETE CASCADE
        )
    """)
    conn.commit()

    cur.execute("SELECT COUNT(*) FROM materias")
    if cur.fetchone()[0] == 0:
        for m in MATERIAS_PADRAO:
            cur.execute("INSERT INTO materias (nome) VALUES (?)", (m,))
        conn.commit()
    conn.close()


# ==========================================================
# UTILITARIOS
# ==========================================================
def limpar_tela():
    os.system("cls" if os.name == "nt" else "clear")


def pausar():
    input("\nPressione ENTER para continuar...")


def hoje():
    return date.today().isoformat()


def data_br(iso):
    if not iso:
        return "-"
    y, m, d = iso.split("-")
    return f"{d}/{m}/{y}"


# ==========================================================
# MATERIAS
# ==========================================================
def listar_materias(conn):
    cur = conn.cursor()
    cur.execute("SELECT id, nome FROM materias ORDER BY nome")
    return cur.fetchall()


def tela_materias(conn):
    while True:
        limpar_tela()
        print("=" * 55)
        print("  MATERIAS")
        print("=" * 55)
        materias = listar_materias(conn)
        for i, (mid, nome) in enumerate(materias, 1):
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM topicos WHERE materia_id=?", (mid,))
            qtd = cur.fetchone()[0]
            print(f" {i:>2}. {nome}  ({qtd} topicos)")
        print("-" * 55)
        print(" [A] Adicionar matéria")
        print(" [R] Remover matéria")
        print(" [0] Voltar ao menu")
        op = input("\nEscolha: ").strip().lower()
        if op == "a":
            nome = input("Nome da nova matéria: ").strip()
            if nome:
                try:
                    conn.execute("INSERT INTO materias (nome) VALUES (?)", (nome,))
                    conn.commit()
                    print("Matéria adicionada!")
                except sqlite3.IntegrityError:
                    print("Essa matéria já existe.")
                pausar()
        elif op == "r":
            idx = input("Número da matéria a remover: ").strip()
            if idx.isdigit() and 1 <= int(idx) <= len(materias):
                mid, nome = materias[int(idx) - 1]
                conf = input(f"Remover '{nome}' e todos os seus tópicos? (s/n): ").lower()
                if conf == "s":
                    conn.execute("DELETE FROM materias WHERE id=?", (mid,))
                    conn.commit()
                    print("Removida.")
                pausar()
        elif op == "0":
            break


# ==========================================================
# TOPICOS
# ==========================================================
def criar_topico(conn, materia_id, nome):
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO topicos (materia_id, nome, data_cadastro, concluido) VALUES (?,?,?,0)",
        (materia_id, nome, hoje())
    )
    topico_id = cur.lastrowid
    for n in range(1, TOTAL_REVISOES + 1):
        cur.execute(
            "INSERT INTO revisoes (topico_id, numero, feita, data_feita) VALUES (?,?,0,NULL)",
            (topico_id, n)
        )
    conn.commit()
    return topico_id


def listar_topicos(conn, materia_id=None):
    cur = conn.cursor()
    if materia_id:
        cur.execute("""
            SELECT t.id, t.nome, m.nome, t.concluido, t.data_cadastro
            FROM topicos t JOIN materias m ON m.id = t.materia_id
            WHERE t.materia_id = ?
            ORDER BY t.id DESC
        """, (materia_id,))
    else:
        cur.execute("""
            SELECT t.id, t.nome, m.nome, t.concluido, t.data_cadastro
            FROM topicos t JOIN materias m ON m.id = t.materia_id
            ORDER BY t.id DESC
        """)
    return cur.fetchall()


def tela_topicos(conn):
    while True:
        limpar_tela()
        print("=" * 55)
        print("  TOPICOS DE ESTUDO")
        print("=" * 55)
        topicos = listar_topicos(conn)
        for i, (tid, nome, materia, concl, dt) in enumerate(topicos, 1):
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM revisoes WHERE topico_id=? AND feita=1", (tid,))
            feitas = cur.fetchone()[0]
            status = "Concluído" if concl else "Em andamento"
            print(f" {i:>2}. [{materia}] {nome}  | {status} | Revisões: {feitas}/{TOTAL_REVISOES}")
        print("-" * 55)
        print(" [N] Novo tópico")
        print(" [E] Editar tópico")
        print(" [C] Marcar/Desmarcar concluído")
        print(" [X] Excluir tópico")
        print(" [0] Voltar ao menu")
        op = input("\nEscolha: ").strip().lower()

        if op == "n":
            materias = listar_materias(conn)
            if not materias:
                print("Cadastre uma matéria primeiro.")
                pausar()
                continue
            print("\nMatérias disponíveis:")
            for i, (mid, nome) in enumerate(materias, 1):
                print(f" {i}. {nome}")
            idx = input("Número da matéria: ").strip()
            if idx.isdigit() and 1 <= int(idx) <= len(materias):
                materia_id = materias[int(idx) - 1][0]
                nome_topico = input("Nome do tópico: ").strip()
                if nome_topico:
                    criar_topico(conn, materia_id, nome_topico)
                    print("Tópico criado com 20 revisões programadas!")
            pausar()

        elif op == "e":
            idx = input("Número do tópico a editar: ").strip()
            if idx.isdigit() and 1 <= int(idx) <= len(topicos):
                tid = topicos[int(idx) - 1][0]
                novo_nome = input("Novo nome: ").strip()
                if novo_nome:
                    conn.execute("UPDATE topicos SET nome=? WHERE id=?", (novo_nome, tid))
                    conn.commit()
                    print("Atualizado!")
            pausar()

        elif op == "c":
            idx = input("Número do tópico: ").strip()
            if idx.isdigit() and 1 <= int(idx) <= len(topicos):
                tid, _, _, concl, _ = topicos[int(idx) - 1]
                conn.execute("UPDATE topicos SET concluido=? WHERE id=?", (0 if concl else 1, tid))
                conn.commit()
            pausar()

        elif op == "x":
            idx = input("Número do tópico a excluir: ").strip()
            if idx.isdigit() and 1 <= int(idx) <= len(topicos):
                tid = topicos[int(idx) - 1][0]
                conf = input("Confirma exclusão? (s/n): ").lower()
                if conf == "s":
                    conn.execute("DELETE FROM topicos WHERE id=?", (tid,))
                    conn.commit()
                    print("Excluído.")
            pausar()

        elif op == "0":
            break


# ==========================================================
# REVISOES
# ==========================================================
def tela_revisoes(conn):
    while True:
        limpar_tela()
        print("=" * 55)
        print("  REVISOES (R1 a R20)")
        print("=" * 55)
        topicos = listar_topicos(conn)
        for i, (tid, nome, materia, concl, dt) in enumerate(topicos, 1):
            cur = conn.cursor()
            cur.execute("SELECT numero, feita FROM revisoes WHERE topico_id=? ORDER BY numero", (tid,))
            revs = cur.fetchall()
            marcado = "".join("X" if f else "." for _, f in revs)
            feitas = sum(1 for _, f in revs if f)
            print(f" {i:>2}. [{materia}] {nome}  ({feitas}/{TOTAL_REVISOES})")
            print(f"      {marcado}")
        print("-" * 55)
        print(" [M] Marcar/Desmarcar uma revisão")
        print(" [0] Voltar ao menu")
        op = input("\nEscolha: ").strip().lower()
        if op == "m":
            idx = input("Número do tópico: ").strip()
            if idx.isdigit() and 1 <= int(idx) <= len(topicos):
                tid = topicos[int(idx) - 1][0]
                n = input(f"Número da revisão (1 a {TOTAL_REVISOES}): ").strip()
                if n.isdigit() and 1 <= int(n) <= TOTAL_REVISOES:
                    cur = conn.cursor()
                    cur.execute("SELECT feita FROM revisoes WHERE topico_id=? AND numero=?", (tid, int(n)))
                    row = cur.fetchone()
                    if row:
                        novo = 0 if row[0] else 1
                        data_feita = hoje() if novo else None
                        conn.execute(
                            "UPDATE revisoes SET feita=?, data_feita=? WHERE topico_id=? AND numero=?",
                            (novo, data_feita, tid, int(n))
                        )
                        conn.commit()
                        print("Atualizado!")
            pausar()
        elif op == "0":
            break


# ==========================================================
# PAINEL
# ==========================================================
def tela_painel(conn):
    limpar_tela()
    print("=" * 55)
    print("  PAINEL DE PROGRESSO")
    print("=" * 55)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM topicos")
    total_topicos = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM topicos WHERE concluido=1")
    concluidos = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM revisoes")
    total_rev = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM revisoes WHERE feita=1")
    feitas_rev = cur.fetchone()[0]
    pct = round((feitas_rev / total_rev) * 100) if total_rev else 0

    print(f" Total de tópicos      : {total_topicos}")
    print(f" Tópicos concluídos    : {concluidos}")
    print(f" Revisões feitas       : {feitas_rev} / {total_rev}")
    print(f" Progresso geral       : {pct}%")
    barra_cheia = int(pct / 5)
    print(" [" + "#" * barra_cheia + "-" * (20 - barra_cheia) + f"] {pct}%")

    print("\n Progresso por matéria:")
    for mid, nome in listar_materias(conn):
        cur.execute("""
            SELECT COUNT(*) FROM revisoes r JOIN topicos t ON t.id=r.topico_id
            WHERE t.materia_id=?
        """, (mid,))
        tot = cur.fetchone()[0]
        cur.execute("""
            SELECT COUNT(*) FROM revisoes r JOIN topicos t ON t.id=r.topico_id
            WHERE t.materia_id=? AND r.feita=1
        """, (mid,))
        fe = cur.fetchone()[0]
        p = round((fe / tot) * 100) if tot else 0
        print(f"   - {nome:<28} {p:>3}%")
    pausar()


# ==========================================================
# EXPORTAR EXCEL
# ==========================================================
def exportar_excel(conn):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("\nBiblioteca 'openpyxl' não encontrada.")
        print("Instale com: pip install openpyxl")
        pausar()
        return

    wb = Workbook()

    titulo_font = Font(bold=True, size=12, color="FFFFFF")
    titulo_fill = PatternFill("solid", fgColor="1C3F9C")
    header_font = Font(bold=True)

    # ---- Aba Materias ----
    ws1 = wb.active
    ws1.title = "Materias"
    ws1.append(["Matéria", "Qtd. Tópicos"])
    for c in ws1[1]:
        c.font = header_font
    for mid, nome in listar_materias(conn):
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM topicos WHERE materia_id=?", (mid,))
        qtd = cur.fetchone()[0]
        ws1.append([nome, qtd])
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 15

    # ---- Aba Topicos ----
    ws2 = wb.create_sheet("Topicos")
    ws2.append(["Matéria", "Tópico", "Status", "Data Cadastro", "Revisões Feitas"])
    for c in ws2[1]:
        c.font = header_font
    for tid, nome, materia, concl, dt in listar_topicos(conn):
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM revisoes WHERE topico_id=? AND feita=1", (tid,))
        feitas = cur.fetchone()[0]
        status = "Concluído" if concl else "Em andamento"
        ws2.append([materia, nome, status, data_br(dt), f"{feitas}/{TOTAL_REVISOES}"])
    for col, w in zip("ABCDE", [26, 34, 16, 15, 16]):
        ws2.column_dimensions[col].width = w

    # ---- Aba Revisoes ----
    ws3 = wb.create_sheet("Revisoes")
    header = ["Matéria", "Tópico"] + [f"R{i}" for i in range(1, TOTAL_REVISOES + 1)]
    ws3.append(header)
    for c in ws3[1]:
        c.font = header_font
    for tid, nome, materia, concl, dt in listar_topicos(conn):
        cur = conn.cursor()
        cur.execute("SELECT numero, feita, data_feita FROM revisoes WHERE topico_id=? ORDER BY numero", (tid,))
        revs = {n: (f, d) for n, f, d in cur.fetchall()}
        linha = [materia, nome]
        for n in range(1, TOTAL_REVISOES + 1):
            f, d = revs.get(n, (0, None))
            linha.append(data_br(d) if f else "-")
        ws3.append(linha)
    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 34

    # ---- Aba Painel ----
    ws4 = wb.create_sheet("Painel de Progresso")
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM topicos")
    total_topicos = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM topicos WHERE concluido=1")
    concluidos = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM revisoes")
    total_rev = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM revisoes WHERE feita=1")
    feitas_rev = cur.fetchone()[0]
    ws4.append(["Indicador", "Valor"])
    for c in ws4[1]:
        c.font = header_font
    ws4.append(["Total de tópicos", total_topicos])
    ws4.append(["Tópicos concluídos", concluidos])
    ws4.append(["Revisões totais", total_rev])
    ws4.append(["Revisões feitas", feitas_rev])
    ws4.append(["Progresso geral (%)", "=IF(B4=0,0,ROUND(B5/B4*100,0))"])
    ws4.column_dimensions["A"].width = 24
    ws4.column_dimensions["B"].width = 16

    nome_arquivo = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Relatorio_Estudos_PMGO_PMDF.xlsx")
    wb.save(nome_arquivo)
    print(f"\nExcel gerado em: {nome_arquivo}")
    pausar()


# ==========================================================
# EXPORTAR PDF
# ==========================================================
def exportar_pdf(conn):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except ImportError:
        print("\nBiblioteca 'reportlab' não encontrada.")
        print("Instale com: pip install reportlab")
        pausar()
        return

    nome_arquivo = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Relatorio_Estudos_PMGO_PMDF.pdf")
    doc = SimpleDocTemplate(nome_arquivo, pagesize=A4,
                             topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Relatório de Estudos - PMGO / PMDF", styles["Title"]))
    story.append(Paragraph(f"Gerado em {data_br(hoje())}", styles["Normal"]))
    story.append(Spacer(1, 14))

    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM topicos")
    total_topicos = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM topicos WHERE concluido=1")
    concluidos = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM revisoes")
    total_rev = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM revisoes WHERE feita=1")
    feitas_rev = cur.fetchone()[0]
    pct = round((feitas_rev / total_rev) * 100) if total_rev else 0

    resumo = [["Tópicos", "Concluídos", "Revisões", "Progresso"],
              [str(total_topicos), str(concluidos), f"{feitas_rev}/{total_rev}", f"{pct}%"]]
    t = Table(resumo, colWidths=[4 * cm] * 4)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C3F9C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    story.append(t)
    story.append(Spacer(1, 18))

    story.append(Paragraph("Tópicos por matéria", styles["Heading2"]))
    dados = [["Matéria", "Tópico", "Status", "Revisões"]]
    for tid, nome, materia, concl, dt in listar_topicos(conn):
        cur.execute("SELECT COUNT(*) FROM revisoes WHERE topico_id=? AND feita=1", (tid,))
        fe = cur.fetchone()[0]
        status = "Concluído" if concl else "Em andamento"
        dados.append([materia, nome, status, f"{fe}/{TOTAL_REVISOES}"])
    if len(dados) == 1:
        dados.append(["-", "Nenhum tópico cadastrado", "-", "-"])
    tabela = Table(dados, colWidths=[4.2 * cm, 6.5 * cm, 3 * cm, 2.3 * cm], repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF2FB")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(tabela)

    doc.build(story)
    print(f"\nPDF gerado em: {nome_arquivo}")
    pausar()


# ==========================================================
# MENU PRINCIPAL
# ==========================================================
def menu_principal():
    iniciar_banco()
    conn = conectar()
    while True:
        limpar_tela()
        print("=" * 55)
        print("   SISTEMA DE ESTUDOS - PMGO / PMDF (Terminal)")
        print("=" * 55)
        print(" 1. Painel de progresso")
        print(" 2. Tópicos de estudo")
        print(" 3. Revisões (R1 a R20)")
        print(" 4. Matérias")
        print(" 5. Exportar Excel")
        print(" 6. Exportar PDF")
        print(" 0. Sair")
        op = input("\nEscolha uma opção: ").strip()

        if op == "1":
            tela_painel(conn)
        elif op == "2":
            tela_topicos(conn)
        elif op == "3":
            tela_revisoes(conn)
        elif op == "4":
            tela_materias(conn)
        elif op == "5":
            exportar_excel(conn)
        elif op == "6":
            exportar_pdf(conn)
        elif op == "0":
            print("Bons estudos! 🚔📘")
            conn.close()
            sys.exit(0)


if __name__ == "__main__":
    try:
        menu_principal()
    except KeyboardInterrupt:
        print("\nSaindo...")
        sys.exit(0)
